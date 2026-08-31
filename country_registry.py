"""Component 1 — Country Registry for HGT-QF.

Maintains the canonical country list used across the whole acquisition system:

    iso3, country_name, region (continent), centroid, bounding box

The bounding box drives the *minimal-subset* spatial requests used by the
scientific data extractor (ERA5/CMIP6). Two provenance levels are tracked:

    bbox_source = "curated"          -> hand-maintained national extent
    bbox_source = "centroid_derived" -> centroid +/- fallback margin (approximate)

The registry can be (re)generated from the research workbook
(data_source_log.xlsx) and the pycountry library, but a static CSV in
config/ is the canonical, version-controlled source of truth.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pycountry

from country_utils import COUNTRY_CENTROIDS, get_country_name, normalize_country

CONFIG_DIR = Path(__file__).parent / "config"
if not (CONFIG_DIR / "country_registry.csv").exists():
    CONFIG_DIR = Path(__file__).parent
COUNTRY_REGISTRY_CSV = CONFIG_DIR / "country_registry.csv"

# ---------------------------------------------------------------------------
# Curated national bounding boxes (north, west, south, east) for the countries
# that matter most for HGT-QF (G20 / EU-27 / Africa Top 12 / Middle East /
# Asia-Pacific / Latin America). All other countries fall back to a
# centroid-derived approximation.
# ---------------------------------------------------------------------------
CURATED_BBOX: dict[str, tuple[float, float, float, float]] = {
    "USA": (49.38, -124.77, 24.40, -66.95), "CAN": (83.11, -141.00, 41.68, -52.62),
    "GBR": (60.85, -8.65, 49.86, 1.77), "DEU": (55.06, 5.87, 47.27, 15.04),
    "FRA": (51.09, -5.14, 41.34, 9.56), "ITA": (47.09, 6.63, 35.49, 18.52),
    "ESP": (43.79, -9.30, 35.94, 4.33), "PRT": (42.15, -9.50, 36.96, -6.19),
    "NLD": (53.55, 3.36, 50.75, 7.23), "BEL": (51.50, 2.55, 49.50, 6.40),
    "CHE": (47.81, 5.96, 45.82, 10.49), "AUT": (49.02, 9.53, 46.37, 17.16),
    "DNK": (57.75, 8.08, 54.56, 15.15), "NOR": (71.18, 4.64, 57.98, 31.06),
    "SWE": (69.06, 11.11, 55.34, 24.16), "FIN": (70.09, 19.09, 59.81, 31.58),
    "POL": (54.84, 14.12, 49.00, 24.15), "CZE": (51.06, 12.09, 48.55, 18.86),
    "SVK": (49.61, 16.84, 47.73, 22.56), "HUN": (48.59, 16.11, 45.74, 22.90),
    "ROU": (48.27, 20.26, 43.62, 29.69), "BGR": (44.22, 22.36, 41.24, 28.61),
    "GRC": (41.75, 19.37, 34.80, 28.25), "IRL": (55.39, -10.66, 51.42, -5.99),
    "RUS": (81.86, 19.25, 41.19, 179.99), "CHN": (53.56, 73.50, 18.16, 134.77),
    "JPN": (45.52, 122.93, 24.05, 145.82), "KOR": (38.61, 124.35, 33.11, 131.87),
    "IND": (35.51, 68.16, 6.75, 97.40), "IDN": (5.90, 95.01, -10.94, 141.02),
    "AUS": (-10.06, 112.92, -43.63, 153.64), "NZL": (-34.39, 165.87, -47.29, 178.55),
    "BRA": (5.27, -73.98, -33.75, -34.79), "MEX": (32.72, -118.36, 14.53, -86.71),
    "ARG": (-21.78, -73.56, -55.06, -53.64), "COL": (13.60, -81.73, -4.23, -66.87),
    "CHL": (-17.50, -75.60, -56.54, -66.42), "PER": (-0.01, -81.33, -18.35, -68.65),
    "VEN": (12.20, -73.35, 0.65, -59.80), "ECU": (1.88, -92.00, -5.01, -75.19),
    "BOL": (-9.68, -69.64, -22.90, -57.45), "URY": (-30.09, -58.49, -34.98, -53.08),
    "PRY": (-19.29, -62.65, -27.61, -54.26), "ZAF": (-22.13, 16.46, -34.83, 32.89),
    "EGY": (31.67, 24.70, 21.72, 36.89), "NGA": (13.89, 2.67, 4.27, 14.68),
    "ETH": (14.89, 33.00, 3.40, 47.98), "KEN": (5.02, 33.91, -4.68, 41.90),
    "GHA": (11.17, -3.26, 4.74, 1.20), "CIV": (10.74, -8.60, 4.36, -2.49),
    "CMR": (13.08, 8.49, 1.65, 16.19), "TZA": (-0.99, 29.33, -11.75, 40.45),
    "DZA": (37.09, -8.67, 18.96, 11.98), "MAR": (35.92, -17.02, 27.67, -0.99),
    "AGO": (-4.39, 11.68, -18.04, 24.08), "SAU": (32.16, 34.49, 16.35, 55.67),
    "ARE": (26.08, 51.58, 22.63, 56.38), "IRN": (39.78, 44.03, 25.06, 63.32),
    "IRQ": (37.38, 38.79, 29.06, 48.57), "JOR": (33.37, 34.96, 29.19, 39.30),
    "KWT": (30.10, 46.55, 28.52, 48.43), "OMN": (26.40, 52.00, 16.65, 59.84),
    "QAT": (26.16, 50.75, 24.48, 51.64), "LBN": (34.69, 35.10, 33.05, 36.64),
    "TUR": (42.11, 25.67, 35.81, 44.83), "ISR": (33.34, 34.27, 29.45, 35.90),
    "UKR": (52.37, 22.14, 44.39, 40.23), "PAK": (37.07, 60.87, 23.69, 77.84),
    "BGD": (26.63, 88.01, 20.74, 92.67), "VNM": (23.39, 102.14, 8.18, 109.46),
    "THA": (20.46, 97.34, 5.61, 105.64), "MYS": (7.36, 99.64, 0.85, 119.27),
    "PHL": (21.12, 116.93, 4.59, 126.60), "SGP": (1.47, 103.61, 1.13, 104.07),
}

# Fallback margin used to derive an approximate bbox from the centroid.
FALLBACK_LAT_MARGIN = 8.0
FALLBACK_LON_MARGIN = 10.0

# Explicit region overrides for countries the workbook does not classify
# cleanly (or omits), plus transcontinental edge cases.
REGION_OVERRIDES: dict[str, str] = {
    "COD": "Africa", "COG": "Africa", "CZE": "Europe", "GNQ": "Africa",
    "MAC": "Asia", "MMR": "Asia", "VAT": "Europe",
}


@dataclass(frozen=True)
class CountryRecord:
    iso3: str
    country_name: str
    region: str
    centroid_lat: float
    centroid_lon: float
    bbox_north: float
    bbox_west: float
    bbox_south: float
    bbox_east: float
    bbox_source: str


def _continent_from_coords(lat: float, lon: float) -> str:
    """Approximate continent from a centroid, used only as a last-resort fallback."""
    if -25 <= lon <= 60 and 34 <= lat <= 72:
        return "Europe"
    if -35 <= lat <= 37 and -18 <= lon <= 52:
        if lat >= 12 and lon >= 34:  # Middle East / Arabian peninsula
            return "Asia"
        return "Africa"
    if -56 <= lat <= 13 and -82 <= lon <= -34:
        return "South America"
    if 5 <= lat <= 85 and -170 <= lon <= -50:
        return "North America"
    if -50 <= lat <= 0 and 110 <= lon <= 180:
        return "Oceania"
    return "Asia"


def _continent_from_pycountry(iso3: str) -> str:
    if iso3 in REGION_OVERRIDES:
        return REGION_OVERRIDES[iso3]
    coords = COUNTRY_CENTROIDS.get(iso3)
    if coords:
        return _continent_from_coords(*coords)
    return "Other"


def _bbox_from_centroid(lat: float, lon: float) -> tuple[float, float, float, float]:
    north = min(90.0, lat + FALLBACK_LAT_MARGIN)
    south = max(-90.0, lat - FALLBACK_LAT_MARGIN)
    west = lon - FALLBACK_LON_MARGIN
    east = lon + FALLBACK_LON_MARGIN
    if west < -180.0:
        west += 360.0
    if east > 180.0:
        east -= 360.0
    return north, west, south, east


def _iter_workbook_countries(workbook_path: Path) -> list[tuple[str, str]]:
    """Yield (iso3, continent) from the 'Global List' sheet of the workbook."""
    out: list[tuple[str, str]] = []
    if not workbook_path.exists():
        return out
    try:
        import openpyxl
    except ImportError:
        return out

    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    if "Global List" not in wb.sheetnames:
        return out
    ws = wb["Global List"]
    continents = {"africa", "asia", "europe", "north america", "south america", "oceania"}
    current = "Other"
    seen: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if s.casefold() in continents:
            current = s.title()
            continue
        iso3 = normalize_country(s)
        if iso3 and iso3 not in seen:
            seen.add(iso3)
            out.append((iso3, current))
    return out


def build_country_records(workbook_path: Path | None = None) -> list[CountryRecord]:
    """Assemble country records from the workbook (or pycountry) + centroids."""
    workbook_path = workbook_path or (Path(__file__).parent / "data_source_log.xlsx")
    source_regions: dict[str, str] = {}
    for iso3, continent in _iter_workbook_countries(workbook_path):
        source_regions[iso3] = continent

    # Baseline: every ISO-3 country with a centroid.
    records: list[CountryRecord] = []
    for iso3 in sorted(COUNTRY_CENTROIDS):
        lat, lon = COUNTRY_CENTROIDS[iso3]
        if iso3 in CURATED_BBOX:
            n, w, s, e = CURATED_BBOX[iso3]
            bbox_source = "curated"
        else:
            n, w, s, e = _bbox_from_centroid(lat, lon)
            bbox_source = "centroid_derived"
        region = source_regions.get(iso3)
        if not region or region.casefold() == "other":
            region = _continent_from_pycountry(iso3)
        records.append(CountryRecord(
            iso3=iso3,
            country_name=get_country_name(iso3),
            region=region,
            centroid_lat=lat,
            centroid_lon=lon,
            bbox_north=round(n, 4),
            bbox_west=round(w, 4),
            bbox_south=round(s, 4),
            bbox_east=round(e, 4),
            bbox_source=bbox_source,
        ))
    return records


def regenerate_country_registry(output_path: Path | None = None) -> Path:
    """(Re)write config/country_registry.csv from current sources."""
    output_path = output_path or COUNTRY_REGISTRY_CSV
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = build_country_records()
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "iso3", "country_name", "region", "centroid_lat", "centroid_lon",
            "bbox_north", "bbox_west", "bbox_south", "bbox_east", "bbox_source",
        ])
        for r in records:
            writer.writerow([
                r.iso3, r.country_name, r.region, r.centroid_lat, r.centroid_lon,
                r.bbox_north, r.bbox_west, r.bbox_south, r.bbox_east, r.bbox_source,
            ])
    return output_path


def load_country_registry() -> dict[str, CountryRecord]:
    """Load the canonical country registry from config/country_registry.csv."""
    registry: dict[str, CountryRecord] = {}
    if not COUNTRY_REGISTRY_CSV.exists():
        return {r.iso3: r for r in build_country_records()}

    with COUNTRY_REGISTRY_CSV.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            iso3 = row["iso3"].strip().upper()
            registry[iso3] = CountryRecord(
                iso3=iso3,
                country_name=row.get("country_name", "").strip(),
                region=row.get("region", "").strip(),
                centroid_lat=float(row.get("centroid_lat", 0) or 0),
                centroid_lon=float(row.get("centroid_lon", 0) or 0),
                bbox_north=float(row.get("bbox_north", 90) or 90),
                bbox_west=float(row.get("bbox_west", -180) or -180),
                bbox_south=float(row.get("bbox_south", -90) or -90),
                bbox_east=float(row.get("bbox_east", 180) or 180),
                bbox_source=row.get("bbox_source", "centroid_derived").strip(),
            )
    return registry


COUNTRY_REGISTRY = load_country_registry()


def get_country_record(iso3: str) -> CountryRecord | None:
    """Return the registry record for an ISO-3 code, or None."""
    return COUNTRY_REGISTRY.get(iso3.strip().upper())


def get_all_countries() -> list[CountryRecord]:
    """Return all registered countries, sorted by ISO-3."""
    return sorted(COUNTRY_REGISTRY.values(), key=lambda r: r.iso3)


def get_country_bbox(iso3: str) -> tuple[float, float, float, float] | None:
    """Return (north, west, south, east) for a country, or None."""
    rec = get_country_record(iso3)
    if rec is None:
        return None
    return rec.bbox_north, rec.bbox_west, rec.bbox_south, rec.bbox_east


def get_countries_in_region(region: str) -> list[str]:
    """Return ISO-3 codes for all countries in a given region/continent."""
    region_norm = region.strip().lower()
    return [
        r.iso3 for r in COUNTRY_REGISTRY.values()
        if r.region.strip().lower() == region_norm
    ]


__all__ = [
    "CountryRecord", "COUNTRY_REGISTRY", "get_country_record", "get_all_countries",
    "get_country_bbox", "get_countries_in_region", "regenerate_country_registry",
    "build_country_records",
]
